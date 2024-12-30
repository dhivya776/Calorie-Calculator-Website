from django.core.management.base import BaseCommand
from django.conf import settings
import os
import openpyxl
from datetime import datetime
from Fityfeed.models import UserProfile  # Import the UserProfile model

class Command(BaseCommand):
    help = 'Download customer details and update the Excel file'

    def handle(self, *args, **kwargs):
        # Get the path where the Excel file will be saved
        file_path = os.path.join(settings.MEDIA_ROOT, 'customer_details.xlsx')

        # Create a workbook and a worksheet if the file doesn't exist
        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Customer Details"
            headers = ["Username", "Email", "Age", "Height", "Weight", "Calorie Needs"]
            for col_num, header in enumerate(headers, start=1):
                sheet.cell(row=1, column=col_num, value=header)
        else:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            # Clear existing data rows (excluding header row)
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                for cell in row:
                    cell.value = None

        # Fetch customer details
        profiles = UserProfile.objects.select_related('user').filter(user__is_staff=False)

        # Write data rows to the sheet
        for row_num, profile in enumerate(profiles, start=2):
            sheet.cell(row=row_num, column=1, value=profile.user.username)
            sheet.cell(row=row_num, column=2, value=profile.user.email)
            sheet.cell(row=row_num, column=3, value=profile.age or "N/A")
            sheet.cell(row=row_num, column=4, value=profile.height or "N/A")
            sheet.cell(row=row_num, column=5, value=profile.weight or "N/A")
            sheet.cell(row=row_num, column=6, value=profile.calorie_needs or "N/A")

        # Save the workbook to the file
        workbook.save(file_path)

        self.stdout.write(self.style.SUCCESS('Successfully updated customer details Excel file.'))

        test_file = os.path.join(settings.MEDIA_ROOT,'test_task_scheduler.txt')
        with open(test_file,'a') as f:
            f.write(f"task executed at {datetime.now()}\n")


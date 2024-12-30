# utils.py (or another suitable name)

import os
import openpyxl
from django.conf import settings
from Fityfeed.models import UserProfile

def update_customer_excel():
    file_path = os.path.join(settings.MEDIA_ROOT, 'customer_details.xlsx')

    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Customer Details"
        headers = ["Username", "Email", "Age", "Height", "Weight", "Calorie Needs"]
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col_num, value=header)
    else:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        # Clear existing data rows
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row:
                cell.value = None

    profiles = UserProfile.objects.select_related('user').filter(user__is_staff=False)
    for row_num, profile in enumerate(profiles, start=2):
        sheet.cell(row=row_num, column=1, value=profile.user.username)
        sheet.cell(row=row_num, column=2, value=profile.user.email)
        sheet.cell(row=row_num, column=3, value=profile.age or "N/A")
        sheet.cell(row=row_num, column=4, value=profile.height or "N/A")
        sheet.cell(row=row_num, column=5, value=profile.weight or "N/A")
        sheet.cell(row=row_num, column=6, value=profile.calorie_needs or "N/A")

    workbook.save(file_path)
